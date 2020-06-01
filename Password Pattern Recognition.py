# -*- coding: utf-8 -*-
"""
Created on Wed Sep 12 12:26:52 2018

@author: Saumya
"""
import cv2
import openpyxl
from tkinter import *
from time import time
from PIL import Image,ImageTk

global j
j=0
global y
y=1
global z
z=0
global l
l=0
global p
p=0
global m
m=0

def wel():
    global roots
    global l
    global m
    global p
    global roott
    global k
    global r1
    global z
    
    if z==1:
        r.destroy()
    
    if l==1:
        roott.destroy()
        
    if p==1:
        r1.destroy()
        
    if m==1:
        roots.destroy()
        
    #print("hello")
    roots = Tk()
    
    k=0
    z=0
    l=0
    p=0
    m=0
    w=800
    h=600
    width,height=roots.winfo_screenwidth(),roots.winfo_screenheight()
    x=(width/2)-(w/2)
    y=(height/2)-(h/2)
    roots.geometry('%dx%d+%d+%d'%(w,h,x,y))
    roots.configure(bg="#90C3C8")
    roots.title('WELCOME')
    
    intruction = Label(roots,font=("Poor Richard",40), text='                                  WELCOME\n',highlightthickness=5,highlightbackground="black",bg="#90C3C8")
    intruction.grid(row=0, column=0,padx=10,pady=2, sticky=N+S)

    loginB = Button(roots, text='Login', command=Login,fg="black",bg="#EA7317",height=3,width=20)
    loginB.place(relx=.41,rely=.3)
    
    signupB = Button(roots, text='Signup', command=Signup,fg="black",bg="#EA7317",height=3,width=20) 
    signupB.place(relx=.41,rely=.5)
    
    analysisB = Button(roots, text='Analysis', command=analysis,fg="black",bg="#EA7317",height=3,width=20)
    analysisB.place(relx=.41,rely=.7)
    
    roots.mainloop()

def Login():
    global pwordE
    global nameE
    global roots
    global k
    global m
    
    if k==0:
        roots.destroy()
    
    roots = Tk()
    m=1
    w=800
    h=600
    width,height=roots.winfo_screenwidth(),roots.winfo_screenheight()
    x=(width/2)-(w/2)
    y=(height/2)-(h/2)
    roots.geometry('%dx%d+%d+%d'%(w,h,x,y))
    roots.configure(bg="#A0CCDA")
    roots.title('LOGIN')
    intruction = Label(roots,font=("Poor Richard",40), text='                                          Login\n',highlightthickness=5,highlightbackground="black",bg="#A0CCDA")
    intruction.grid(row=0, column=0,padx=10,pady=2, sticky=N+S)


    nameL = Label(roots,font=("",20), text='Username   :: ',highlightthickness=5,bg="#A0CCDA")
    nameL.place(relx=.2,rely=.25)

    nameE = Entry(roots,bg="lightgrey",highlightthickness=2,highlightbackground="grey")
    nameE.place(relx=0.5,rely=0.28)

    loginB = Button(roots, text='confirm', command=gett,fg="black",bg="#EA7317",height=3,width=20)
    loginB.place(relx=.42,rely=.6)
    
    backB = Button(roots, text='Back', command=wel,fg="black",bg="#EA7317",height=2,width=10)
    backB.place(relx=.6,rely=.9)
    
    roots.mainloop()
    
def gett():
    global rot
    global roots
    global counter
    global chk
    global cmp
    global totaltime
    rot=Tk()
    w=800
    h=600
    width,height=rot.winfo_screenwidth(),rot.winfo_screenheight()
    x=(width/2)-(w/2)
    y=(height/2)-(h/2)
    rot.geometry('%dx%d+%d+%d'%(w,h,x,y))
    rot.configure(bg="#A0CCDA")

    mylist = []
    file = 'data.xlsx'
    wb = openpyxl.load_workbook(filename=file)
    ws = wb["Sheet1"]
    rows=ws.max_row
    for i in range(1,rows+1):
        mylist.extend([ws.cell(row=i,column=1).value])

    pwordL = Label(rot,font=("",20), text='Password   :: ',highlightthickness=5,bg="#A0CCDA")
    pwordL.place(relx=.2,rely=.25)
    pwordE = Entry(rot,bg="lightgrey",highlightthickness=2,highlightbackground="grey",show='*')
    pwordE.place(relx=0.5,rely=0.28)
    tw=pwordE.get()
    name=nameE.get()

    if name in mylist:
        print="bye"
        counter=mylist.index(name)
        counter+=1
        cmp=ws.cell(row=counter,column=2).value

        roots.destroy()

        loginB = Button(rot, text='Login', command=CheckLogin,fg="black",bg="#EA7317",height=3,width=20)
        loginB.place(relx=.42,rely=.6)
        def logg(keyevent):
            global totaltime
            cword = cmp
            cwordsize = len(cword)
            cwordlist = tuple(cword)
            cwordfl = str(cwordlist[0])
            cwordll = str(cwordlist[-1])

            tword = pwordE.get()
            twordsize = len(tword)
            if twordsize > 0:
                twordlist = tuple(tword)
            twordfl = str(twordlist[0])
            twordll = str(twordlist[-1])
            if cwordsize == 1 and twordsize == 1:
                print("more letters")
            if twordsize == 1 and cwordsize > 1:
                global start
                start = time()
            if twordsize == cwordsize and twordsize != 1:
                if cwordll == twordll:
                    stop = time()
                    totaltime = stop - start
            
            new_col=[tword]
            file='data.xlsx'
            wb = openpyxl.load_workbook(filename=file)
            ws = wb["Sheet1"]
            col=11
            for row, entry in enumerate(new_col, start=1):
                ws.cell(row=counter, column=col, value=entry)

            wb.save(file)
            wb.close()

        pwordE.bind('<KeyRelease>', logg)

def CheckLogin():
    ##############################
    global cmp
    global totaltime
    rot.destroy()
    file = 'data.xlsx'
    wb = openpyxl.load_workbook(filename=file)
    ws = wb["Sheet1"]
    tw=ws.cell(row=counter,column=11).value
    t1=ws.cell(row=counter,column=9).value
    t2=ws.cell(row=counter,column=10).value
    print("time",totaltime)
    if tw == cmp and t1<=totaltime and totaltime<=t2:
        global z
        z=1
        global m
        m=0
        
        global r
        global user
        file = 'data.xlsx'
        wb = openpyxl.load_workbook(filename=file)
        ws = wb["Sheet1"]
        user=ws.cell(row=counter,column=1).value
        
        r = Tk()
        w=700
        h=550
        width,height=r.winfo_screenwidth(),r.winfo_screenheight()
        x=(width/2)-(w/2)
        y=(height/2)-(h/2)
        r.geometry('%dx%d+%d+%d'%(w,h,x,y))
        r.configure(bg="#A0CCDA")
        r.title(':D')
        
        rlbl = Label(r,font=("Poor Richard",30),text='\n Hello {}'.format(user),bg="#A0CCDA")
        rlbl.pack()
       
        analysisB = Button(r, text='Log Out', command=wel,fg="black",bg="#EA7317",height=3,width=20)
        analysisB.place(relx=.6,rely=.85)
        
        loginB = Button(r, text='CheckIntruders',command=CheckIntruder, fg="black",bg="#EA7317",height=3,width=20)
        loginB.place(relx=.4,rely=.4)
        
        r.mainloop()
    else:
        global p
    
        
        p=1
        
        file = 'data.xlsx'
        wb = openpyxl.load_workbook(filename=file)
        ws = wb["Sheet1"]
        use=ws.cell(row=counter,column=1).value
        
        cam = cv2.VideoCapture(0)       
        ret, frame = cam.read()
                        
        img_name = "pic_{}.png".format(use)
        cv2.imwrite(img_name, frame)
        cam.release()

        global r1
        r1 = Tk()
        w=600
        h=450
        width,height=r1.winfo_screenwidth(),r1.winfo_screenheight()
        x=(width/2)-(w/2)
        y=(height/2)-(h/2)
        r1.geometry('%dx%d+%d+%d'%(w,h,x,y))
        r1.configure(bg="#A0CCDA")
        r1.title(':D')
        
        rlbl = Label(r1,font=("Poor Richard",40),text='\nInvalid \n Login',bg="#A0CCDA")
        rlbl.pack()
        
        logoutB = Button(r1, text='Exit', command=close,fg="black",bg="#EA7317",height=3,width=20)
        logoutB.place(relx=.6,rely=.85)
        
        r1.mainloop()

def close():
    global r1
    r1.destroy()
def display():
    global root 
    root.destroy()
    
    global r
    global m
    m=0
    global z
    z=1
    global user
    
    file = 'data.xlsx'
    wb = openpyxl.load_workbook(filename=file)
    ws = wb["Sheet1"]
    user=ws.cell(row=counter,column=1).value
    
    r = Tk()
    w=700
    h=550
    width,height=r.winfo_screenwidth(),r.winfo_screenheight()
    x=(width/2)-(w/2)
    y=(height/2)-(h/2)
    r.geometry('%dx%d+%d+%d'%(w,h,x,y))
    r.configure(bg="#A0CCDA")
    r.title(':D')
    
    rlbl = Label(r,font=("Poor Richard",30),text='\n[+] Logged In',bg="#A0CCDA")
    rlbl.pack()
    
    analysisB = Button(r, text='Log Out', command=wel,fg="black",bg="#EA7317",height=3,width=20)
    analysisB.place(relx=.6,rely=.85)
    
    loginB = Button(r, text='CheckIntruders',command=CheckIntruder, fg="black",bg="#EA7317",height=3,width=20)
    loginB.place(relx=.4,rely=.4)
    
    r.mainloop()
def CheckIntruder():
    global r
    r.destroy()
    
    global root
    root=Tk()
    frame1 = Frame(root)
    frame1.pack(side=TOP, fill=X)
    
    photo1 = PhotoImage(file="pic_{}.png".format(user))
    
    button1 = Button(frame1, compound=TOP, image=photo1,font=("",17),text="Back",fg="white", bg='Black', command=display)
    button1.pack()
    
    button1.image = photo1
    
    root.mainloop()

def Signup():
    global nameEL
    global pwordEL
    global emailEL
    global mobEL
    global ageEL
    global rootA
    global roots
    global rows

    roots.destroy()
    fields = 'Username', 'Password', 'Email', 'Mob','Age'
    def fetch(entries):
        file = 'data.xlsx'
        wb = openpyxl.load_workbook(filename=file)
        ws = wb["Sheet1"]
        rows=ws.max_row
        rows+=1

        col=0

        for entry in entries:
            text  = entry[1].get()

            new_col = [text]

            col+=1

            for row, entry in enumerate(new_col, start=1):
                ws.cell(row=rows, column=col, value=entry)

            wb.save(file)

    def makeform(root, fields):
       entries = []
       for field in fields:
          row = Frame(rootA)
          lab = Label(row, width=35,font=("",13), text=field+"  ::", anchor='w',bg="#90C3C8")
          ent = Entry(row)
          row.pack(side=TOP, fill=Y, padx=5, pady=5)
          lab.pack(side=LEFT)
          ent.pack(side=RIGHT, expand=YES, fill=Y)
          entries.append((field, ent))
       return entries

    if __name__ == '__main__':
       fields = 'Username', 'Password', 'Email', 'Mob','Age'
       rootA = Tk()
       w=800
       h=600
       width,height=rootA.winfo_screenwidth(),rootA.winfo_screenheight()
       x=(width/2)-(w/2)
       y=(height/2)-(h/2)
       rootA.geometry('%dx%d+%d+%d'%(w,h,x,y))
       rootA.configure(bg="#90C3C8")

       rootA.title('SIGN-UP')
       intruction = Label(rootA,font=("",20), text='Please Enter  Credentials\n',bg="#90C3C8")
       intruction.pack(side=LEFT, padx=5, pady=5)
       ents = makeform(rootA, fields)
       rootA.bind('<Return>', (lambda event, e=ents: fetch(e)))

       b2 = Button(rootA, text='Next',command=combine_funcs((lambda e=ents: fetch(e)), sign),fg="black",bg="#EA7317",height=3,width=10)
       b2.place(relx=0.42,rely=0.8)

       rootA.mainloop()

def combine_funcs(*funcs):
    def combined_func(*args, **kwargs):
        for f in funcs:
            f(*args, **kwargs)
    return combined_func

def sign():
     rootA.destroy()
     pas()

def DelUser():
    rootA.destroy()
    Login()
def paes():
    global roots
    roots.destroy()
    pas()
def exi():
    global roots
    roots.destroy()
    Login()

def pas():
    global j
    j+=1
    global roots

    roots = Tk()
    w=700
    h=500
    width,height=roots.winfo_screenwidth(),roots.winfo_screenheight()
    x=(width/2)-(w/2)
    y=(height/2)-(h/2)
    roots.geometry('%dx%d+%d+%d'%(w,h,x,y))
    roots.configure(bg="#A0CCDA")
    roots.title('LOGIN')
    
    intruction = Label(roots,font=("Poor Richard",30), text='          CONFIRM     PASSWORD \n',highlightthickness=5,highlightbackground="black",bg="#A0CCDA")
    intruction.grid(row=0, column=0,padx=10,pady=2, sticky=N+S)
    
    w1 = Label(roots,font=("",20), text='Password   :: ',highlightthickness=5,bg="#A0CCDA")    
    w1.grid(row=3, column=0,padx=10,pady=2, sticky=N+S)   
    
    we = Entry(roots,bg="lightgrey",highlightthickness=2,highlightbackground="grey")
    we.grid(row=3, column=1,padx=10,pady=2, sticky=N+S)
    
    wb = openpyxl.load_workbook('data.xlsx')
    ws = wb['Sheet1']
    for row in ws.iter_rows('B{}:B{}'.format(ws.min_row,ws.max_row)):
        for cell in row:
            continue
        ce=cell.value
    sheet = wb.active

    def pressed(keyevent):
        cword = ce
        cwordsize = len(cword)
        cwordlist = tuple(cword)
        cwordfl = str(cwordlist[0])
        cwordll = str(cwordlist[-1])

        tword = we.get()
        twordsize = len(tword)
        if twordsize > 0:
            twordlist = tuple(tword)
        twordfl = str(twordlist[0])
        twordll = str(twordlist[-1])
        if cwordsize == 1 and twordsize == 1:
            print("more letters")
        if twordsize == 1 and cwordsize > 1:
            global start
            start = time()
        if twordsize == cwordsize and twordsize != 1:
            if cwordll == twordll:
                stop = time()
                totaltime = stop - start
                print("time=",totaltime)
                file = 'data.xlsx'
                new_col = [totaltime]

                wb = openpyxl.load_workbook(filename=file)
                ws = wb["Sheet1"]
                rows = ws.max_row
                if j==1:
                    col = 6
                if j==2:
                    col = 7
                if j==3:
                    col = 8


                for row, entry in enumerate(new_col, start=1):
                    ws.cell(row=rows, column=col, value=entry)

                wb.save(file)
                wb.close()

    we.bind('<KeyRelease>', pressed)
    
    if j<3:
        okButton = Button(roots, text='Confirm', command=paes,fg="black",bg="#EA7317",height=3,width=20)
        okButton.place(relx=.42,rely=.6)

    else:
       ekButton = Button(roots, text='End', command=cal,fg="black",bg="#EA7317",height=3,width=20)
       ekButton.place(relx=.42,rely=.6)

def cal():
    global k
    j=0
    if j==0:
        global roots
    
        wb = openpyxl.load_workbook('data.xlsx')
        ws = wb['Sheet1']
        rows=ws.max_row
        sum=0
        sheet = wb.active
        for col in ws.iter_cols(min_row=rows,max_row=rows,min_col=6,max_col=8):
            for cell in col:
                sum=sum+cell.value
    
        avg=sum/3
        minavg=avg-0.3
        max_avg=avg+0.3
        print("min=",minavg)
        print("max=",max_avg)
        new_col = [minavg]
        file='data.xlsx'
        wb = openpyxl.load_workbook(filename=file)
        ws = wb["Sheet1"]
        col=9
        for row, entry in enumerate(new_col, start=1):
            ws.cell(row=rows, column=col, value=entry)
        new_col = [max_avg]
        col=10
        for row, entry in enumerate(new_col, start=1):
            ws.cell(row=rows, column=col, value=entry)
    
        wb.save(file)
        wb.close()
    
        j=1
    if j==1:
        k=1
        roots.destroy()
        Login()


def analysis():
    global roots 
    roots.destroy()
    global roott
    global l
    roott = Tk()
    l=1
    
    frame1 = Frame(roott)
    frame1.pack(side=TOP, fill=X)
    
    photo1 = PhotoImage(file="analyy.png")
    
    button1 = Button(frame1, compound=BOTTOM, image=photo1,font=("",17),text="Back", bg='green', command=wel)
    button1.pack()

    button1.image = photo1
    
    roott.mainloop()
 

wel()
